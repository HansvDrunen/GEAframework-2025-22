# Import libraries
import os, re, time, random
import io
import fitz  # PyMuPDF
from sentence_transformers import SentenceTransformer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from transformers import pipeline
import numpy as np
import pandas as pd
import streamlit as st
import requests
from urllib.parse import urlparse, urljoin, urlsplit, urlunsplit, urlunparse
from urllib import robotparser
from bs4 import BeautifulSoup
import tldextract
from duckduckgo_search import DDGS
from textwrap import dedent
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ---------------------------
# Streamlit page (MUST be first Streamlit command)
# ---------------------------
st.set_page_config(
    page_title="GEA Statements Extractor",
    layout="wide",
    initial_sidebar_state="expanded"
)

# -----------------------------
# CSS (app-like sidebar + cards)
# -----------------------------
st.markdown(
    dedent("""
    <style>
    /* =========================
       General layout
       ========================= */
    .block-container {
        padding-top: 1.2rem;
        padding-bottom: 2.5rem;
        max-width: 1400px;
    }

    /* =========================
       Sidebar
       ========================= */
    [data-testid="stSidebar"] {
        background: linear-gradient(
            180deg,
            rgba(255,255,255,0.04),
            rgba(255,255,255,0.02)
        );
        border-right: 1px solid rgba(255,255,255,0.08);
    }

    [data-testid="stSidebar"] .block-container {
        padding-top: 1.2rem;
    }

    .side-title {
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 20px;
        font-weight: 900;
        margin: 0.2rem 0 0.6rem 0;
        letter-spacing: -0.02em;
    }

    .side-sub {
        color: rgba(255,255,255,0.72);
        font-size: 13px;
        margin-top: -6px;
        margin-bottom: 1rem;
    }

    /* =========================
       Top header / hero
       ========================= */
    .topbar {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 0.95rem 1.05rem;
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 18px;
        background: rgba(255,255,255,0.03);
        backdrop-filter: blur(6px);
        margin-bottom: 1.25rem;
    }

    .topbar-center {
        justify-content: center;
        text-align: center;
    }

    .hero {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 6px;
    }

    .hero-title {
        font-size: 36px;
        font-weight: 950;
        letter-spacing: -0.03em;
        line-height: 1.05;
    }

    .hero-sub {
        color: rgba(255,255,255,0.72);
        font-size: 15px;
        max-width: 860px;
        text-align: center;
    }

    .hero-badge {
        font-size: 12px;
        padding: 7px 12px;
        border-radius: 999px;
        border: 1px solid rgba(255,255,255,0.10);
        background: rgba(255,255,255,0.04);
        color: rgba(255,255,255,0.78);
        margin-top: 6px;
        white-space: nowrap;
    }

    /* =========================
       Cards
       ========================= */
    .card {
        border-radius: 18px;
        padding: 1.1rem;
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(255,255,255,0.08);
    }

    /* Fix HTML spacing inside cards */
    .card h3 {
        margin-top: 0;
        margin-bottom: 0.5rem;
    }

    .card p {
        margin: 0.4rem 0;
        line-height: 1.65;
    }

    .card ul {
        margin: 0.5rem 0 0 1.1rem;
    }

    .card li {
        margin: 0.25rem 0;
        line-height: 1.6;
    }

    /* =========================
       Text helpers
       ========================= */
    .muted {
        color: rgba(255,255,255,0.72);
    }

    .small {
        font-size: 13px;
        color: rgba(255,255,255,0.72);
    }

    /* =========================
       Pills
       ========================= */
    .pillrow {
        display: flex;
        flex-wrap: wrap;
        gap: 8px;
        margin-top: 0.6rem;
    }

    .pill {
        font-size: 12px;
        padding: 6px 10px;
        border-radius: 999px;
        border: 1px solid rgba(255,255,255,0.10);
        background: rgba(255,255,255,0.04);
        color: rgba(255,255,255,0.80);
    }

    /* =========================
       Responsive
       ========================= */
    @media (max-width: 900px) {
        .hero-title { font-size: 28px; }
        .hero-sub { font-size: 14px; }
    }
    </style>
    """),
    unsafe_allow_html=True
)

# ---------- Sidebar nav ----------
with st.sidebar:
    st.markdown('<div class="side-title">🔎 GEA Extractor</div>', unsafe_allow_html=True)
    st.markdown('<div class="side-sub">Mission / Vision / Values / Strategy / Goals</div>', unsafe_allow_html=True)

    page = st.radio(
        "Navigation",
        ["Home", "Webscrapper", "PDF Extractor", "Coherence"],
        index=0,
        label_visibility="collapsed",
    )

    st.divider()
    st.markdown("**Quick info**")
    st.caption("Runs your pipeline: Discovery → candidates → scoring → best_sites_top.csv")

# ---------- Top header (main) ----------
st.markdown(
    dedent("""
    <div class="topbar topbar-center">
      <div class="hero">
        <div class="hero-title">GEA Statements Extractor</div>
        <div class="hero-sub">Extract Mission / Vision / Values / Strategy / Goals from Web & PDFs</div>
        <div class="hero-badge">Discovery → candidates → scoring → export</div>
      </div>
    </div>
    """),
    unsafe_allow_html=True
)

# ---------- Home summary page ----------
def render_home():
    # IMPORTANT: NO dedent + NO leading indentation in the HTML strings
    # (otherwise Streamlit may render it like a code block)

    html_1 = """
<div class="card">
  <h3>What this app does</h3>

  <div class="muted" style="line-height: 1.6;">
    This demo extracts enterprise statements from <b>official web pages</b> and <b>PDF reports</b>.
    It then structures the results into a clean table so you can compare statements across companies
    and (later) measure coherence.
  </div>

  <div class="pillrow" style="margin-top: 14px;">
    <span class="pill">Mission</span>
    <span class="pill">Vision</span>
    <span class="pill">Core Values</span>
    <span class="pill">Strategy</span>
    <span class="pill">Goals</span>
  </div>

  <div style="height: 14px;"></div>

  <div class="small">
    <b>Flow</b>: discovery → scoring → extraction → export → coherence (next)
  </div>
</div>
""".strip()

    html_2 = """
<div class="card">
  <h3>Modules</h3>

  <p class="muted">
    Use the sidebar to switch modules. Each module is designed to produce an exportable file for analysis.
  </p>

  <div style="height: 14px;"></div>

  <div class="muted">

    <p><b>🌐 Webscrapper</b></p>
    <ul>
      <li>Finds the company’s official website</li>
      <li>Discovers relevant pages (mission, values, sustainability, about)</li>
      <li>Ranks pages using keyword signals + metadata (title, headings)</li>
    </ul>
    <p class="small"><b>Outputs:</b> candidates.csv, manifest.csv, best_sites_top.csv</p>

    <div style="height: 16px;"></div>

    <p><b>📄 PDF Extractor</b></p>
    <ul>
      <li>Upload a PDF report (e.g., annual report)</li>
      <li>Extracts text per page and splits into sentences</li>
      <li>Classifies statements into GEA categories</li>
      <li>Exports a final table for analysis</li>
    </ul>
    <p class="small"><b>Outputs:</b> gea_final_table.csv, gea_final_table.xlsx</p>

    <div style="height: 16px;"></div>

    <p><b>📏 Coherence</b></p>
    <ul>
      <li>Upload reviewed statements (Excel)</li>
      <li>Create a coherence matrix with human-in-the-loop scoring</li>
      <li>Export the matrix to Excel</li>
    </ul>
    <p class="small"><b>Outputs:</b> coherence_matrix.xlsx</p>

  </div>
</div>
""".strip()

    html_3 = """
<div class="card">
  <h3>How to use</h3>

  <div class="muted" style="line-height: 1.7;">
    <b>1)</b> Open <b>Webscrapper</b> → run <b>Discovery</b> to collect candidate pages → (optional) run <b>Scoring</b> to rank the best official pages.<br>
    <b>2)</b> Open <b>PDF Extractor</b> → upload a report → run extraction + classification → download the <b>final_table</b> (Excel).<br>
    <b>3)</b> Open <b>Coherence</b> → compare alignment across statements using element extraction + matrix scoring → export Excel.
  </div>
</div>
""".strip()

    st.markdown(html_1, unsafe_allow_html=True)
    st.write("")
    st.markdown(html_2, unsafe_allow_html=True)
    st.write("")
    st.markdown(html_3, unsafe_allow_html=True)

# ---------------------------
# Add cached model + cached query embeddings
# ---------------------------
@st.cache_resource
def load_embedder():
    return SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")

# Keep these SHORT (long academic definitions reduce cosine scores)
QUERIES = {
  "Mission": "Definition: A mission is a brief, typically one sentence, statement that defines the fundamental purpose of an enterprise that is'enduringly pursued but never fulfilled'.", 
  "Vision": "Definition: An enterprise vision is a concise statement that operationalises a mission, it is external and market oriented and should express, preferably in aspirational terms, how the enterprise intends to be perceived by the world",
  "Core Values": "Definition: A core value statement for an enterprise prescribes a set of desired behaviours, character and culture and is required for an enterprise to be, or become, successful within its formulated vision. Core values are used to indicate the value platforms from which the enterprise wants to operate and, in the most fundamental sense, they give substance and direction to what is important for an enterprise. We consider core values to be guiding statements for enterprises held at the highest level of purpose. Core values can also be referred to as statements of the fundamental principles of an enterprise. It is what management considers most important for the enterprises functioning, and considers critical for behaviour within an enterprise. Core values are not easily changed.",  
  "Strategy": "Definition: A strategy statement of an enterprise is a form of comprehensive master plan in which it is stated how an enterprise will achieve its goals. A strategy statement outlines the way an enterprise wants to achieve its goals, at which the 'whats, mission, vision, core values, goals, of the enterprise are translated into the 'how. How will the enterprise achieve its goals? This strategy-building process is generally carried out continuously within an enterprise, both for the shorter, medium and longer term development of the enterprise.",
  "Goals": "A goal statement is a formulation of a desired stage of development for an enterprise working towards achieving the enterprise's vision. Goals are a set of ambitions that an enterprise has, translated into short, medium and long-term goals. The term goals at this level is used to indicate the collective ambitions, for example, customer focus, continuity, sustainability, profit growth and internationalisation. Goals at the level of purpose of the enterprise are also referred to in the literature as 'goals. To make a distinction: objectives on a more concrete level can be referred to as 'targets' or 'objectives. This creates a goal hierarchy with, at the top the mission, vision and core values and the goals determined within those frameworks, followed by the objectives derived from this."
}

@st.cache_resource
def load_query_embeddings():
    model = load_embedder()
    labels = list(QUERIES.keys())
    emb = model.encode(list(QUERIES.values()), normalize_embeddings=True)
    return labels, emb

# ---------------------------
# Add your sentence splitter + rule override
# ---------------------------
def split_sentences(text: str):
    if not isinstance(text, str):
        return []
    t = text.strip()
    if not t:
        return []
    parts = re.split(r'(?<=[.!?])\s+', t)
    return [p.strip() for p in parts if p and p.strip()]

def rule_override(text, current_label):
    t = (text or "").lower()

    # CORE VALUES
    if any(w in t for w in [
        "values", "value", "kernwaarden", "waarde", "integrity", "respect", "ethics",
        "principles", "principes", "inclusive", "inclusion", "committed", "commitment"
    ]):
        return "Core Values"

    # MISSION
    if any(w in t for w in [
        "mission", "missie", "purpose", "doel", "why we exist", "exist to", "raison d"
    ]):
        return "Mission"

    # VISION
    if any(w in t for w in [
        "vision", "visie", "future", "toekomst", "aspire", "ambition", "ambitie", "aim to become", "we want to be"
    ]):
        return "Vision"

    # STRATEGY
    if any(w in t for w in [
        "strategy", "strategic", "strategie", "strategisch",
        "our approach", "aanpak", "we focus on", "focus", "roadmap", "priorities", "prioriteiten"
    ]):
        return "Strategy"

    # GOALS
    if any(w in t for w in [
        "goals", "goal", "doelen", "doel", "objective", "objectives", "target", "targets",
        "aim to", "we aim", "we plan", "we will", "by 20", "in 20"
    ]):
        return "Goals"

    return current_label


# ---------------------------
# PDF text extraction (per page)
# ---------------------------
def extract_pages_from_pdf(file_bytes: bytes, max_pages: int = 80):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages = []
    n = min(len(doc), max_pages)
    for i in range(n):
        text = doc.load_page(i).get_text("text")
        text = re.sub(r"\s+", " ", text).strip()
        if text:
            pages.append({"page": i + 1, "text": text})
    doc.close()
    return pages


# ---------------------------
# Sentence-level classification (fast)
# ---------------------------
def classify_sentences(sentences):
    model = load_embedder()
    q_labels, q_emb = load_query_embeddings()

    sent_emb = model.encode(sentences, normalize_embeddings=True)
    sims = cosine_similarity(sent_emb, q_emb)
    best_idx = sims.argmax(axis=1)
    best_conf = sims.max(axis=1)

    preds = []
    for s, i, c in zip(sentences, best_idx, best_conf):
        lab = q_labels[i]
        preds.append((s, lab, float(c)))
    return preds


# ---------------------------
# Build final table
# ---------------------------
def build_final_table_from_pdf(pdf_name: str, pages, conf_sent=0.20, max_sents_per_page=120):
    rows = []
    chunk_id = 0

    for p in pages:
        chunk_id += 1
        sents = split_sentences(p["text"])
        if not sents:
            continue

        # cap to keep laptop fast
        sents = sents[:max_sents_per_page]

        preds = classify_sentences(sents)

        for sid, (sent, pred, conf) in enumerate(preds, start=1):
            if conf < conf_sent:
                continue

            final_lab = rule_override(sent, pred)

            rows.append({
                "pdf": pdf_name,
                "page": p["page"],
                "source_chunk_id": chunk_id,
                "statement": sent,
                "gea_category": final_lab
            })

    return pd.DataFrame(rows)


# ---------------------------
# Defaults / settings
# ---------------------------
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 (+streamlit-demo)")
TIMEOUT = 20

# --------- Discovery config (same as your notebook) ----------
KEYWORDS_URL = [
    "mission","vision","purpose","values","about","who-we-are","our-company",
    "strategy","culture","sustainability","corporate-governance","what-we-stand-for",
    "purpose-and-values","purpose-values","at-a-glance","principles","code-of-conduct"
]
CANDIDATE_PATHS = [
    "/", "/about", "/about-us", "/company", "/who-we-are", "/our-company",
    "/values", "/purpose", "/mission", "/vision", "/about/mission", "/about/vision",
    "/sustainability", "/culture", "/our-values", "/strategy", "/purpose-and-values"
]

def is_probably_official(domain: str, company_name: str) -> bool:
    bad = {
        "linkedin.com","facebook.com","instagram.com","x.com","twitter.com",
        "wikipedia.org","crunchbase.com","glassdoor.com","bloomberg.com",
        "reuters.com","yahoo.com","google.com","news.google.com","youtube.com"
    }
    ext = tldextract.extract(domain)
    root = f"{ext.domain}.{ext.suffix}" if ext.suffix else ext.domain
    if root in bad:
        return False
    tokens = re.findall(r"[a-z0-9]+", company_name.lower())
    hits = sum(tok in ext.domain.lower() for tok in tokens if len(tok) >= 3)
    return hits >= 1

def normalize_home(url: str) -> str:
    url = url.strip().rstrip("/")
    if not url.startswith("http"):
        url = "https://" + url
    return url

def ddg_with_backoff(query, max_results=10, attempts=3):
    delay = 1.5
    for _ in range(attempts):
        try:
            with DDGS(timeout=TIMEOUT) as ddgs:
                results = list(ddgs.text(query, max_results=max_results, region="wt-wt", safesearch="off"))
                if results:
                    return results
        except Exception:
            time.sleep(delay + random.random())
            delay *= 2
    return []

def find_official_homepage(company_name: str, override_homepage: str) -> str|None:
    if override_homepage.strip():
        return normalize_home(override_homepage)

    results = ddg_with_backoff(f"{company_name} official site", max_results=10, attempts=3)
    for r in results:
        url = r.get("href") or r.get("url")
        if not url:
            continue
        p = urlparse(url)
        if p.scheme.startswith("http"):
            base = f"{p.scheme}://{p.netloc}"
            if is_probably_official(base, company_name):
                return base.rstrip("/")
    return None

def allowed_by_robots(url: str) -> bool:
    try:
        p = urlparse(url)
        base = f"{p.scheme}://{p.netloc}"
        rp = robotparser.RobotFileParser()
        rp.set_url(urljoin(base, "/robots.txt"))
        rp.read()
        return rp.can_fetch(UA, url)
    except Exception:
        return True

def fetch(url: str):
    return requests.get(
        url,
        headers={"User-Agent": UA, "Accept-Language": "en"},
        timeout=TIMEOUT,
        allow_redirects=True
    )

def cleanup_url(u: str) -> str:
    if not u:
        return u
    u = u.strip()
    if not u:
        return u
    if not u.startswith(("http://","https://")):
        u = "https://" + u
    p = urlsplit(u)
    return urlunsplit((p.scheme, p.netloc, p.path, "", ""))

def ensure_htm(url: str) -> str:
    try:
        p = urlparse(url)
        path_lower = p.path.lower()
        if p.netloc.endswith("ing.com") and "/about-us/" in path_lower:
            last = p.path.rstrip("/").split("/")[-1]
            if last and ("." not in last) and re.search(r"[a-zA-Z\-]", last):
                return url + ".htm"
    except Exception:
        pass
    return url

def preflight(url: str) -> tuple[bool,str,int]:
    try:
        r = requests.head(url, headers={"User-Agent": UA}, allow_redirects=True, timeout=12)
        if 200 <= r.status_code < 300:
            return True, r.url, r.status_code
        if r.status_code in (403, 405):
            rg = requests.get(url, headers={"User-Agent": UA}, allow_redirects=True, timeout=12)
            return (200 <= rg.status_code < 300), rg.url, rg.status_code
        return False, getattr(r, "url", url), r.status_code
    except Exception:
        try:
            rg = requests.get(url, headers={"User-Agent": UA}, allow_redirects=True, timeout=12)
            return (200 <= rg.status_code < 300), rg.url, rg.status_code
        except Exception:
            return False, url, 0

def discover_candidates(base: str, max_candidates=60):
    seen, cands = set(), []

    def add(u, reason, boost=0.0):
        u = u.split("#")[0].rstrip("/")
        if urlparse(u).netloc != urlparse(base).netloc:
            return
        if u in seen:
            return
        seen.add(u)
        cands.append({"url": u, "reason": reason, "boost": boost})

    for p in CANDIDATE_PATHS:
        add(urljoin(base, p), "seed", 0.2 if p != "/" else 0.0)

    # Sitemap scan
    for sm in ["/sitemap.xml", "/sitemap_index.xml", "/sitemap-index.xml"]:
        sm_url = urljoin(base, sm)
        if not allowed_by_robots(sm_url):
            continue
        try:
            r = fetch(sm_url)
            if not (r.ok and "xml" in (r.headers.get("content-type",""))):
                continue
            soup = BeautifulSoup(r.text, "xml")
            locs = [loc.get_text(strip=True) for loc in soup.find_all("loc")]
            # follow some child sitemaps
            child_maps = [u for u in locs if u.endswith(".xml")]
            for cm in child_maps[:10]:
                try:
                    rr = fetch(cm)
                    if rr.ok and "xml" in (rr.headers.get("content-type","")):
                        s2 = BeautifulSoup(rr.text, "xml")
                        locs.extend([loc.get_text(strip=True) for loc in s2.find_all("loc")])
                except:
                    pass
            for u in locs:
                if any(k in u.lower() for k in KEYWORDS_URL):
                    add(u, "sitemap", 0.9)
        except:
            pass

    # DDG site search
    try:
        with DDGS(timeout=TIMEOUT) as ddgs:
            query = f"site:{urlparse(base).netloc} " + " ".join(KEYWORDS_URL[:6])
            for r in ddgs.text(query, max_results=25, region="wt-wt", safesearch="off"):
                u = r.get("href") or r.get("url")
                if u:
                    add(u, "site_search", 1.0)
    except:
        pass

    return cands[:max_candidates]

def is_hub_path(path: str) -> bool:
    pl = path.lower()
    return any(pl == x or pl.startswith(x + "/") for x in [
        "/about-us", "/sustainability", "/about", "/company", "/who-we-are", "/our-company"
    ])

def extract_samehost_links(base_url: str, html: str, limit=120):
    base_host = urlparse(base_url).netloc
    soup = BeautifulSoup(html, "lxml")
    out = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("#"):
            continue
        absu = urljoin(base_url, href)
        pu = urlparse(absu)
        if pu.netloc != base_host:
            continue
        low = absu.lower()
        if any(k in low for k in KEYWORDS_URL):
            out.append(absu.split("#")[0])
        if len(out) >= limit:
            break
    # de-dupe
    seen, ded = set(), []
    for u in out:
        if u not in seen:
            seen.add(u)
            ded.append(u)
    return ded

def score_discovery(url: str, reason: str, boost: float) -> float:
    path = urlparse(url).path
    low = path.lower()
    s = 0.0
    for k in KEYWORDS_URL:
        if k in low:
            s += 1.25
    s += boost
    if low.endswith(".htm"):
        s += 0.8
    s += max(0, 2.0 - 0.12*len(path))
    if path.count("/") <= 2:
        s += 0.3
    if any(x in low for x in ["/about","/purpose","/values","/mission","/vision","/strategy","/sustainability"]):
        s += 0.5
    if reason == "site_search":
        s += 0.3
    if reason == "page_links":
        s += 0.6
    return round(s, 3)

# --------- Scoring pipeline (your 2nd script) ----------
KEYWORD_WEIGHTS = {
    "mission": 10, "vision": 10, "purpose": 9, "values": 9,
    "strategy": 8, "sustainability": 10, "esg": 9, "csr": 8,
    "about": 7, "who-we-are": 7, "our-company": 7, "company": 5,
    "annual": 9, "report": 8, "integrated": 7, "governance": 6,
    "code-of-conduct": 6, "principles": 6, "culture": 6, "impact": 6,
    "responsibility": 6, "investor": 6, "purpose-and-values": 9,
}
LABEL_RULES = [
    ("mission", "Mission"),
    ("vision", "Vision"),
    ("purpose", "Purpose"),
    ("values", "Values"),
    ("strategy", "Strategy"),
    ("sustainability|esg|csr|responsibility|impact", "Sustainability"),
    (r"\bannual\b|\breport\b|\bintegrated\b|\binvestor\b", "Annual/Report"),
    ("about|who-we-are|our-story|our-company|company", "About/Company"),
    ("governance|code-of-conduct|principles", "Governance"),
]
LABEL_PRIORITY = [
    "Mission","Vision","Values","Purpose","Strategy",
    "Sustainability","Annual/Report","About/Company","Governance","General"
]
LABEL_RANK = {lab: i for i, lab in enumerate(LABEL_PRIORITY)}

class RobotsCache:
    def __init__(self):
        self._cache = {}
    def allowed(self, user_agent: str, url: str) -> bool:
        netloc = urlparse(url).netloc
        if not netloc:
            return False
        rp = self._cache.get(netloc)
        if rp is None:
            rp = robotparser.RobotFileParser()
            robots_url = f"https://{netloc}/robots.txt"
            try:
                rp.set_url(robots_url)
                rp.read()
            except Exception:
                rp = None
            self._cache[netloc] = rp
        if self._cache[netloc] is None:
            return True
        try:
            return self._cache[netloc].can_fetch(user_agent, url)
        except Exception:
            return True

ROBOTS = RobotsCache()

def normalize_url(url: str) -> str:
    if not url:
        return url
    p = urlparse(url)
    scheme = p.scheme or "https"
    return urlunparse((scheme, p.netloc, p.path or "/", p.params, p.query, ""))

def strip_noncontent(soup: BeautifulSoup) -> None:
    for sel in [
        "nav","header","footer","[role='banner']","[role='navigation']","[role='contentinfo']",
        ".cookie","#cookie","[id*='cookie']","[class*='cookie']",
        ".consent","[id*='consent']",".newsletter",".subscribe",".social",".share",
        "script","style","noscript","svg"
    ]:
        for el in soup.select(sel):
            el.decompose()

def _score_blob(text: str) -> int:
    text = (text or "").lower()
    return sum(w for kw, w in KEYWORD_WEIGHTS.items() if kw in text)

def _label_for(url: str, title: str, h1: str, h2: str) -> str:
    blob = " ".join(x for x in [url, title, h1, h2] if x).lower()
    for pat, lab in LABEL_RULES:
        if re.search(pat, blob):
            return lab
    return "General"

def score_page(url: str, html: str):
    soup = BeautifulSoup(html, "lxml")
    strip_noncontent(soup)
    title = soup.title.get_text(" ", strip=True) if soup.title else ""
    h1_el, h2_el = soup.select_one("h1"), soup.select_one("h2")
    h1 = h1_el.get_text(" ", strip=True) if h1_el else ""
    h2 = h2_el.get_text(" ", strip=True) if h2_el else ""
    score = 2*_score_blob(url) + 2*_score_blob(title) + _score_blob(h1) + _score_blob(h2)
    return {"url": url, "title": title, "h1": h1, "h2": h2, "label": _label_for(url, title, h1, h2), "score": score}

def retry_fetch(url: str, attempts=3, backoff=1.5):
    last_exc = None
    for i in range(attempts):
        try:
            r = fetch(url)
            if 200 <= r.status_code < 400:
                return r
            last_exc = RuntimeError(f"HTTP {r.status_code}")
        except Exception as e:
            last_exc = e
        time.sleep(backoff * (2 ** i))
    raise last_exc if last_exc else RuntimeError("Unknown fetch error")

def company_from_url(u: str) -> str:
    try:
        netloc = urlparse(u).netloc.lower()
        for prefix in ("www.", "corporate.", "ir.", "investors.", "about.", "en."):
            if netloc.startswith(prefix):
                netloc = netloc[len(prefix):]
        parts = netloc.split(".")
        brand = parts[0] if len(parts) >= 2 else netloc
        brand = brand.replace("-", " ").strip()
        return brand.upper() if 2 <= len(brand) <= 4 else brand.title()
    except Exception:
        return "(unknown)"

def run_scoring_from_candidates(candidates_df: pd.DataFrame, top_per_company: int = 5):
    # expects a URL column named "final_url" or "url"
    if "final_url" in candidates_df.columns:
        urls = candidates_df["final_url"].dropna().astype(str).tolist()
    else:
        urls = candidates_df["url"].dropna().astype(str).tolist()

    targets = [{"url": normalize_url(u), "company": company_from_url(u)} for u in urls]

    rows_scored, rows_manifest = [], []
    for item in targets:
        url, company = item["url"], item["company"]
        rowm = {"company": company, "url": url, "status": "", "http": "", "ctype": "", "label": "", "score": "", "title": "", "h1": "", "h2": ""}

        try:
            if not ROBOTS.allowed(UA, url):
                rowm["status"] = "robots_blocked"
                rows_manifest.append(rowm)
                continue

            resp = retry_fetch(url)
            ctype = resp.headers.get("content-type","").lower()
            rowm["http"] = getattr(resp, "status_code", "")
            rowm["ctype"] = ctype

            if "html" not in ctype:
                rowm["status"] = "non_html"
                rows_manifest.append(rowm)
                continue

            scored = score_page(resp.url, resp.text)
            scored["company"] = company
            rows_scored.append(scored)

            rowm.update({
                "status": "ok",
                "label": scored["label"],
                "score": scored["score"],
                "title": scored["title"],
                "h1": scored["h1"],
                "h2": scored["h2"],
                "url": resp.url
            })
            rows_manifest.append(rowm)

        except Exception as e:
            rowm["status"] = f"error:{type(e).__name__}"
            rows_manifest.append(rowm)

    man_df = pd.DataFrame(rows_manifest, columns=["company","url","status","http","ctype","label","score","title","h1","h2"])
    raw_df = pd.DataFrame(rows_scored).drop_duplicates(["company","url"])

    if raw_df.empty:
        return man_df, raw_df, raw_df

    raw_df["label_rank"] = raw_df["label"].map(LABEL_RANK).fillna(len(LABEL_RANK))
    ranked = raw_df.sort_values(["company","label_rank","score"], ascending=[True, True, False])
    best = ranked.groupby("company").head(top_per_company)
    return man_df, raw_df, best

def extract_clean_text(url: str, max_chars: int = 12000) -> str:
    r = fetch(url)
    soup = BeautifulSoup(r.text, "lxml")
    for tag in soup(["script","style","noscript","header","footer","nav","aside"]):
        tag.decompose()
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"\n{2,}", "\n\n", text)
    return text[:max_chars]


# ============================================================
# UI WRAPPERS (USE SIDEBAR `page` FROM YOUR CSS PART)
# ============================================================

def render_webscrapper():
    st.subheader("🌐 Webscrapper")

    col1, col2, col3 = st.columns([1.2, 2, 1])
    with col1:
        company = st.text_input("Company name", "ING")
    with col2:
        override_home = st.text_input("Homepage override", "https://ing.com/")
    with col3:
        max_candidates = st.number_input("MAX_CANDIDATES", min_value=10, max_value=300, value=60, step=10)

    with st.expander("Hub expansion"):
        expand_from_hubs = st.toggle("EXPAND_FROM_HUBS", value=True)
        max_hubs = st.slider("MAX_HUBS_TO_EXPAND", 0, 15, 5, 1)
        max_links_per_hub = st.slider("MAX_LINKS_PER_HUB", 20, 500, 120, 10)

    with st.expander("Scoring"):
        top_per_company = st.slider("TOP_PER_COMPANY", 1, 10, 5, 1)

    c1, c2 = st.columns([1, 1])
    with c1:
        run_discovery = st.button("🚀 Run Discovery", use_container_width=True)
    with c2:
        run_scoring = st.button("⭐ Run Scoring (from discovery results)", use_container_width=True)

    # ---------------------------
    # Discovery run
    # ---------------------------
    if run_discovery:
        base = find_official_homepage(company, override_home)
        if not base:
            st.error("Could not determine official homepage. Provide a correct homepage override.")
            st.stop()

        st.success(f"Using homepage: {base}")
        raw = discover_candidates(base, max_candidates=int(max_candidates))

        if expand_from_hubs:
            hub_candidates = [c for c in raw if is_hub_path(urlparse(c["url"]).path)][:max_hubs]
            expanded = []
            for hc in hub_candidates:
                hub_url = cleanup_url(hc["url"])
                if not allowed_by_robots(hub_url):
                    continue
                try:
                    r = fetch(hub_url)
                    if r.ok and "html" in (r.headers.get("content-type","") or "").lower():
                        links = extract_samehost_links(r.url, r.text, limit=max_links_per_hub)
                        for u in links:
                            expanded.append({"url": u, "reason": "page_links", "boost": 1.0})
                except:
                    pass
            raw.extend(expanded)

        seen_urls = set()
        rows = []
        for c in raw:
            u0 = c["url"]
            u1 = ensure_htm(cleanup_url(u0))
            if u1 in seen_urls:
                continue
            seen_urls.add(u1)
            ok, final_u, status = preflight(u1)
            rows.append({
                "url": u0,
                "normalized_url": u1,
                "final_url": final_u,
                "reason": c["reason"],
                "score": score_discovery(u1, c["reason"], c["boost"]),
                "status_code": status,
                "is_live": bool(ok)
            })

        df = pd.DataFrame(rows)
        df.sort_values(by=["is_live","score"], ascending=[False, False], inplace=True)

        st.session_state["candidates_df"] = df

        st.subheader("✅ candidates.csv")
        st.dataframe(df, use_container_width=True)

        st.download_button(
            "⬇️ Download candidates.csv",
            df.to_csv(index=False).encode("utf-8"),
            "candidates.csv",
            "text/csv"
        )

        live_urls = df[df["is_live"]].head(17)["final_url"].tolist()
        st.download_button(
            "⬇️ Download top17_live.txt",
            "\n".join(live_urls).encode("utf-8"),
            "top17_live.txt",
            "text/plain"
        )

    # ---------------------------
    # Scoring run
    # ---------------------------
    if run_scoring:
        if "candidates_df" not in st.session_state:
            st.warning("Run Discovery first (or load a candidates.csv).")
            st.stop()

        cand = st.session_state["candidates_df"]
        live = cand[cand["is_live"]].copy()

        if live.empty:
            st.warning("No live URLs to score.")
            st.stop()

        with st.spinner("Scoring pages (fetching HTML + computing labels/scores)…"):
            man_df, raw_df, best_df = run_scoring_from_candidates(live, top_per_company=int(top_per_company))

        st.subheader("🧾 manifest.csv")
        st.dataframe(man_df, use_container_width=True)
        st.download_button(
            "⬇️ Download manifest.csv",
            man_df.to_csv(index=False).encode("utf-8"),
            "manifest.csv",
            "text/csv"
        )

        st.subheader("⭐ best_sites_top.csv")
        if best_df.empty:
            st.warning("No HTML pages were scored (maybe blocked by robots / non-HTML).")
            st.stop()

        st.dataframe(best_df, use_container_width=True)
        st.download_button(
            "⬇️ Download best_sites_top.csv",
            best_df.to_csv(index=False).encode("utf-8"),
            "best_sites_top.csv",
            "text/csv"
        )

        st.session_state["best_df"] = best_df

    # ---------------------------
    # Preview section
    # ---------------------------
    st.divider()
    st.subheader("🔍 Preview a selected page")

    best_df = st.session_state.get("best_df", None)
    cand_df = st.session_state.get("candidates_df", None)

    options = []
    if best_df is not None and not best_df.empty:
        options = best_df["url"].dropna().astype(str).tolist()
    elif cand_df is not None and not cand_df.empty:
        options = cand_df[cand_df["is_live"]]["final_url"].dropna().astype(str).tolist()

    if options:
        sel = st.selectbox("Pick a URL", options)
        if st.button("Load page text"):
            with st.spinner("Fetching & cleaning text…"):
                try:
                    st.text_area("Extracted text (truncated)", extract_clean_text(sel), height=380)
                except Exception as e:
                    st.error(f"Failed to fetch text: {e}")
    else:
        st.info("Run Discovery (and optionally Scoring) to enable preview.")

def render_pdf_extractor():
    st.subheader("📄 PDF Extraction")

    # --- Fixed defaults (no UI sliders) ---
    MIN_CHARS = 10
    MAX_CHARS = 300
    MAX_PER_CAT = 25
    PAGE_SIZE = 10
    gea_keep_order = ["Mission", "Vision", "Strategy", "Core Values", "Goals"]

    # --- Upload ---
    up = st.file_uploader("Upload a PDF", type=["pdf"], key="pdf_uploader")

    colA, colB = st.columns([1, 1])
    with colA:
        CONF_SENT = st.slider("Confidence threshold", 0.05, 0.35, 0.20, 0.01)
    with colB:
        MAX_PAGES = st.number_input("Max pages", min_value=1, max_value=500, value=80, step=10)

    run_btn = st.button("🚀 Run PDF extraction + classification", use_container_width=True)

    # ------------------------------------------------------------
    # Reset session if a new file is uploaded
    # ------------------------------------------------------------
    if "last_pdf_name" not in st.session_state:
        st.session_state.last_pdf_name = None

    if up is not None and up.name != st.session_state.last_pdf_name:
        st.session_state.last_pdf_name = up.name
        st.session_state.pop("pdf_final_table_review", None)
        st.session_state.pop("review_page", None)

    # ------------------------------------------------------------
    # If we already have results in session, show review UI directly
    # ------------------------------------------------------------
    if "pdf_final_table_review" in st.session_state and st.session_state.pdf_final_table_review is not None:
        df = st.session_state.pdf_final_table_review
        if "review_page" not in st.session_state:
            st.session_state.review_page = 0

        # ---------- Pagination controls ----------
        total = len(df)
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)

        navL, navC, navR = st.columns([1, 2, 1])
        with navL:
            if st.button("⬅️ Prev", use_container_width=True, disabled=(st.session_state.review_page <= 0)):
                st.session_state.review_page -= 1
        with navC:
            st.markdown(
                f"<div style='text-align:center; padding-top:10px;'>"
                f"<b>Page {st.session_state.review_page+1} / {total_pages}</b> — showing {PAGE_SIZE} statements</div>",
                unsafe_allow_html=True,
            )
        with navR:
            if st.button("Next ➡️", use_container_width=True, disabled=(st.session_state.review_page >= total_pages - 1)):
                st.session_state.review_page += 1

        st.caption("Edits are saved automatically. Use Prev/Next to review more statements.")

        # ---------- Page slice ----------
        start = st.session_state.review_page * PAGE_SIZE
        end = min(start + PAGE_SIZE, total)

        page_df = df.iloc[start:end].copy()

        # Show a nice editor table
        st.markdown("### 🧠 Human review (interactive table)")

        edited = st.data_editor(
            page_df[["keep", "statement", "review_category", "page"]].rename(
                columns={"review_category": "gea_category"}
            ),
            use_container_width=True,
            hide_index=True,
            column_config={
                "keep": st.column_config.CheckboxColumn("Keep", help="Select if you want to keep this statement."),
                "statement": st.column_config.TextColumn("Statement", disabled=True, width="large"),
                "gea_category": st.column_config.SelectboxColumn("GEA category", options=gea_keep_order),
                "page": st.column_config.NumberColumn("Page", disabled=True, width="small"),
            },
            disabled=["statement", "page"],
            key=f"editor_page_{st.session_state.review_page}",
        )

        # Write edited results back into the master df
        # (align by the original row index of the slice)
        slice_index = df.iloc[start:end].index
        df.loc[slice_index, "keep"] = edited["keep"].values
        df.loc[slice_index, "review_category"] = edited["gea_category"].values

        st.session_state.pdf_final_table_review = df

        # ---------- Selected statements table ----------
        st.markdown("### ✅ Selected statements (live table)")
        reviewed_df = df[df["keep"]].copy()
        reviewed_df["gea_category"] = reviewed_df["review_category"]

        st.dataframe(
            reviewed_df[["statement", "gea_category", "page"]].reset_index(drop=True),
            use_container_width=True,
            height=260,
        )

        # ---------- Export ----------
        if reviewed_df.empty:
            st.warning("No statements selected yet. Tick Keep to include statements in the export.")
            return

        st.success(f"✔ {len(reviewed_df)} statements selected")

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            reviewed_df[["statement", "gea_category", "page"]].to_excel(
                writer, index=False, sheet_name="reviewed_statements"
            )
        buffer.seek(0)

        st.download_button(
            "⬇️ Download reviewed statements (Excel)",
            data=buffer.getvalue(),
            file_name="gea_reviewed_statements.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        return

    # ------------------------------------------------------------
    # Otherwise run extraction/classification
    # ------------------------------------------------------------
    if not run_btn:
        return

    if up is None:
        st.warning("Please upload a PDF first.")
        st.stop()

    with st.spinner("Extracting text from PDF pages…"):
        file_bytes = up.read()
        pages = extract_pages_from_pdf(file_bytes, max_pages=int(MAX_PAGES))

    if not pages:
        st.error("No text found in the PDF (it might be scanned images).")
        st.stop()

    st.success(f"Extracted text from {len(pages)} pages.")

    with st.spinner("Classifying statements (multilingual embeddings)…"):
        final_table = build_final_table_from_pdf(
            pdf_name=up.name,
            pages=pages,
            conf_sent=float(CONF_SENT),
            # max_sents_per_page is controlled in your pipeline already
        )

    if final_table is None or final_table.empty:
        st.warning("No statements passed the confidence threshold. Try lowering it (e.g. 0.15).")
        st.stop()

    # ------------------------------------------------------------
    # CLEAN/REDUCE (fixed: MIN_CHARS=30, MAX_PER_CAT=25)
    # ------------------------------------------------------------
    final_table = final_table.copy()

    # Normalize whitespace
    final_table["statement_clean"] = (
        final_table["statement"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
    )

    # Drop empty
    final_table = final_table[final_table["statement_clean"].str.len() > 0].copy()

    # Drop duplicates
    final_table = final_table.drop_duplicates(subset=["statement_clean", "gea_category"]).copy()

    # Keep only human-readable statements (length-controlled)
    final_table = final_table[
    (final_table["statement_clean"].str.len() >= MIN_CHARS) &
    (final_table["statement_clean"].str.len() <= MAX_CHARS)
    ].copy()


    # Limit per category: keep the longest statements
    final_table["_len"] = final_table["statement_clean"].str.len()

    reduced_parts = []
    for cat in gea_keep_order:
        sub = (
            final_table[final_table["gea_category"] == cat]
            .sort_values("_len", ascending=False)
            .head(MAX_PER_CAT)
        )
        reduced_parts.append(sub)

    final_table = pd.concat(reduced_parts, ignore_index=True) if reduced_parts else final_table.iloc[0:0].copy()

    final_table = final_table.drop(columns=["statement_clean", "_len"], errors="ignore")

    if final_table.empty:
        st.warning("After cleaning/reducing, no statements remain. Try lowering confidence threshold.")
        st.stop()

    # ------------------------------------------------------------
    # Prepare review table in session state (interactive + paginated)
    # ------------------------------------------------------------
    review_df = final_table.copy()

    # ensure these columns exist
    if "page" not in review_df.columns:
        review_df["page"] = None

    review_df["keep"] = True
    review_df["review_category"] = review_df["gea_category"].fillna("").astype(str).str.strip()

    # Map to valid categories if something weird is inside
    review_df.loc[~review_df["review_category"].isin(gea_keep_order), "review_category"] = "Mission"

    st.session_state.pdf_final_table_review = review_df
    st.session_state.review_page = 0

    st.success(f"Done. Prepared {len(review_df)} statements for review (paged 10 per page).")

    # Immediately show review UI on the same run
    st.rerun()



def render_coherence():
    st.subheader("📏 Measure coherence")

    st.write(
        "Upload the Excel you exported from the PDF Extractor (reviewed statements), "
        "choose a relationship, extract elements (human-in-the-loop), score the matrix "
        "(human-in-the-loop with optional auto-suggestions), and export to Excel."
    )

    # ---------------------------
    # 0) Upload
    # ---------------------------
    up = st.file_uploader("Upload reviewed statements Excel (.xlsx)", type=["xlsx"])

    if up is None:
        st.info("Upload a reviewed statements Excel file to begin.")
        return

    try:
        df = pd.read_excel(up)
    except Exception as e:
        st.error(f"Could not read Excel: {e}")
        return

    # ---- Normalize column names (handles gea-category vs gea_category, casing, spaces) ----
    df.columns = [str(c).strip().lower().replace("-", "_").replace(" ", "_") for c in df.columns]

    required = {"statement", "gea_category"}
    if not required.issubset(set(df.columns)):
        st.error("Your Excel must contain columns: 'statement' and 'gea_category'.")
        st.write("Columns found:", df.columns.tolist())
        return

    # ---- Normalize categories to canonical labels ----
    def normalize_cat(x):
        x = "" if pd.isna(x) else str(x).strip().lower()
        mapping = {
            "mission": "Mission",
            "vision": "Vision",
            "strategy": "Strategy",
            "strategies": "Strategy",
            "core value": "Core Values",
            "core values": "Core Values",
            "values": "Core Values",
            "goal": "Goals",
            "goals": "Goals",
        }
        return mapping.get(x, x.title())

    df["gea_category"] = df["gea_category"].apply(normalize_cat)
    df["statement"] = df["statement"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()

    # Preview
    st.dataframe(df.head(50), use_container_width=True, height=280)

    # ---------------------------
    # 1) Choose relationship + caps
    # ---------------------------
    st.markdown("## 1) Choose a relationship to score")

    relation = st.selectbox(
        "Relationship",
        ["Mission ↔ Vision", "Strategy ↔ Goals"],
        index=0
    )

    if relation == "Mission ↔ Vision":
        # caps: mission fixed at 1, vision 1–6
        max_mission = 1
        max_vision = st.number_input("Max statements from Vision", min_value=1, max_value=6, value=4, step=1)
        row_label = "Vision statements"
        col_label = "Mission elements"
    else:
        # caps: strategy 1–10, goals 1–5
        max_strategy = st.number_input("Max statements from Strategy", min_value=1, max_value=10, value=7, step=1)
        max_goals = st.number_input("Max statements from Goals", min_value=1, max_value=5, value=3, step=1)
        row_label = "Strategy statements"
        col_label = "Goals elements"

    # ---------------------------
    # 2) Select statements (table)
    # ---------------------------
    st.markdown("## 2) Select statements to use (table)")

    if relation == "Mission ↔ Vision":
        mission_df = df[df["gea_category"] == "Mission"].head(max_mission)
        vision_df = df[df["gea_category"] == "Vision"].head(int(max_vision))

        if mission_df.empty or vision_df.empty:
            st.warning("You need at least 1 Mission and 1 Vision statement in the uploaded file.")
            return

        chosen_A = mission_df["statement"].tolist()   # Mission statements (used only for extracting mission elements)
        chosen_B = vision_df["statement"].tolist()    # Vision statements (matrix columns)

        st.caption(f"Using {len(chosen_A)} Mission statement(s) and {len(chosen_B)} Vision statement(s).")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Mission (selected)**")
            st.dataframe(mission_df[["statement", "gea_category"]], use_container_width=True, height=220)
        with col2:
            st.markdown("**Vision (selected)**")
            st.dataframe(vision_df[["statement", "gea_category"]], use_container_width=True, height=220)

    else:
        strategy_df = df[df["gea_category"] == "Strategy"].head(int(max_strategy))
        goals_df = df[df["gea_category"] == "Goals"].head(int(max_goals))

        if strategy_df.empty or goals_df.empty:
            st.warning("You need at least 1 Strategy and 1 Goals statement in the uploaded file.")
            return

        chosen_A = strategy_df["statement"].tolist()  # Strategy statements (matrix rows)
        chosen_B = goals_df["statement"].tolist()     # Goals statements (used only for extracting goal elements)

        st.caption(f"Using {len(chosen_A)} Strategy statement(s) and {len(chosen_B)} Goals statement(s).")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Strategy (selected)**")
            st.dataframe(strategy_df[["statement", "gea_category"]], use_container_width=True, height=220)
        with col2:
            st.markdown("**Goals (selected)**")
            st.dataframe(goals_df[["statement", "gea_category"]], use_container_width=True, height=220)

    # ---------------------------
    # 3) Element extraction + HUMAN review (only for one side)
    # ---------------------------
    st.markdown("## 3) Extract elements (human-in-the-loop)")

    # ---- Simple element extraction (frequency bigrams/unigrams) ----
    STOP = set("""
        a an the and or but if then else when while to of in on for with by from as at into over under
        is are was were be been being this that these those it its their our your we you they i
        company companies group bank banking
        """.split())

    def _tokens(text):
        text = re.sub(r"[^a-zA-Z0-9\s]", " ", text.lower())
        words = [w for w in text.split() if len(w) >= 3 and w not in STOP]
        return words

    def extract_elements_from_statements(statements, top_k=12):
        # Build unigram + bigram frequencies
        freq = {}
        for s in statements:
            words = _tokens(s)
            for w in words:
                freq[w] = freq.get(w, 0) + 1
            for i in range(len(words) - 1):
                bg = f"{words[i]} {words[i+1]}"
                freq[bg] = freq.get(bg, 0) + 1

        # sort by frequency then length (prefer more informative phrases)
        items = sorted(freq.items(), key=lambda x: (x[1], len(x[0])), reverse=True)
        out = []
        seen = set()
        for phrase, _ in items:
            if phrase in seen:
                continue
            seen.add(phrase)
            out.append(phrase)
            if len(out) >= top_k:
                break
        return out

    def editable_elements_block(title, base_key, initial_elements):
        """
        Human-in-the-loop editor:
          - checkboxes to include/exclude
          - text inputs to edit
          - add new element
          - confirm -> returns selected list
        """
        st.markdown(f"### {title}")

        # Initialize state
        if f"{base_key}_elements" not in st.session_state:
            st.session_state[f"{base_key}_elements"] = list(initial_elements)

        if f"{base_key}_use" not in st.session_state:
            st.session_state[f"{base_key}_use"] = [True] * len(st.session_state[f"{base_key}_elements"])

        # Make lengths match
        elems = st.session_state[f"{base_key}_elements"]
        use_flags = st.session_state[f"{base_key}_use"]
        if len(use_flags) != len(elems):
            use_flags = [True] * len(elems)
            st.session_state[f"{base_key}_use"] = use_flags

        rows = []
        for i, el in enumerate(elems):
            c1, c2 = st.columns([0.18, 0.82])
            with c1:
                use_flags[i] = st.checkbox("Use", value=use_flags[i], key=f"{base_key}_use_{i}")
            with c2:
                new_val = st.text_input("", value=str(el), key=f"{base_key}_txt_{i}")
                elems[i] = new_val.strip()

        st.session_state[f"{base_key}_elements"] = elems
        st.session_state[f"{base_key}_use"] = use_flags

        # Add new
        cA, cB = st.columns([0.75, 0.25])
        with cA:
            new_el = st.text_input("Add new:", value="", placeholder="Type new element", key=f"{base_key}_add")
        with cB:
            if st.button("Add", key=f"{base_key}_add_btn"):
                if new_el.strip():
                    st.session_state[f"{base_key}_elements"].append(new_el.strip())
                    st.session_state[f"{base_key}_use"].append(True)
                    st.session_state[f"{base_key}_add"] = ""

        if st.button("Confirm elements", key=f"{base_key}_confirm"):
            pass  # just forces rerun nicely

        selected = [
            e for e, use in zip(st.session_state[f"{base_key}_elements"], st.session_state[f"{base_key}_use"])
            if use and str(e).strip()
        ]
        return selected

    # Extract only one side depending on relation
    if relation == "Mission ↔ Vision":
        # only mission elements (vision stays as full statements)
        auto_mission_elements = extract_elements_from_statements(chosen_A, top_k=14)
        mission_elements = editable_elements_block(
            "Mission elements (editable)",
            base_key="mv_mission",
            initial_elements=auto_mission_elements
        )

        # For MV:
        # rows = mission elements
        # cols = vision statements
        rows = mission_elements
        cols = chosen_B

        if len(rows) == 0:
            st.warning("No mission elements selected. Tick 'Use' and/or add a new element.")
            return

    else:
        # Strategy ↔ Goals: only goal elements (strategy stays as statements)
        auto_goal_elements = extract_elements_from_statements(chosen_B, top_k=14)
        goal_elements = editable_elements_block(
            "Goals elements (editable)",
            base_key="sg_goals",
            initial_elements=auto_goal_elements
        )

        # rows = strategy statements
        # cols = goal elements
        rows = chosen_A
        cols = goal_elements

        if len(cols) == 0:
            st.warning("No goal elements selected. Tick 'Use' and/or add a new element.")
            return

    # ---------------------------
    # 4) Matrix scoring (auto-suggest + human override)
    # ---------------------------
    st.markdown("## 4) Score the coherence matrix (human-in-the-loop)")
    st.caption("Scale: -3 (contradicts) … 0 (no clear relation) … +3 (strongly supports).")

    # ---- Embedding + optional NLI (as you requested) ----
    @st.cache_resource
    def _load_embedder():
        return SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")

    def embed(texts):
        model = _load_embedder()
        return model.encode(list(texts), normalize_embeddings=True)

    @st.cache_resource
    def _load_nli_pipeline():
        try:
            return pipeline("text-classification", model="roberta-large-mnli", top_k=None)
        except Exception:
            return None

    def nli_probs(premises, hypotheses):
        nli = _load_nli_pipeline()
        if nli is None:
            z = np.zeros(len(premises), dtype=float)
            return z, z, z

        contras, neutrals, entails = [], [], []
        for prem, hyp in zip(premises, hypotheses):
            pair = f"{prem} </s></s> {hyp}"
            out = nli(pair)
            d = {x["label"].upper(): float(x["score"]) for x in out}
            contras.append(d.get("CONTRADICTION", 0.0))
            neutrals.append(d.get("NEUTRAL", 0.0))
            entails.append(d.get("ENTAILMENT", 0.0))
        return np.array(contras), np.array(neutrals), np.array(entails)

    def score_pair(strategy, goal):
        # semantic relatedness (main driver)
        s_vec = embed([strategy])[0]
        g_vec = embed([goal])[0]
        sim = float(np.dot(s_vec, g_vec))

        # NLI only for negativity (conflict detector)
        hypothesis = f"The company's goal is: {goal}."
        c, n, e = nli_probs([strategy], [hypothesis])
        c = float(c[0]); n = float(n[0]); e = float(e[0])

        # Strong contradiction => negative
        if c >= 0.60 and sim >= 0.20:
            return -2, sim, c, n, e
        if c >= 0.40 and sim >= 0.20:
            return -1, sim, c, n, e

        # Otherwise, score by similarity (support strength)
        if sim >= 0.55:
            return 3, sim, c, n, e
        if sim >= 0.42:
            return 2, sim, c, n, e
        if sim >= 0.22:
            return 1, sim, c, n, e

        return 0, sim, c, n, e

    def auto_score_matrix(rows, cols):
        scores = pd.DataFrame(0, index=rows, columns=cols)
        debug_rows = []
        for r in rows:
            for c in cols:
                sc, sim, contr, neu, ent = score_pair(r, c)
                scores.loc[r, c] = int(sc)
                debug_rows.append({
                    "row": r,
                    "col": c,
                    "score": int(sc),
                    "sim": sim,
                    "contradiction": contr,
                    "neutral": neu,
                    "entailment": ent,
                })
        return scores, pd.DataFrame(debug_rows)

    # ---- Initialize / keep matrix in session_state ----
    state_key = f"matrix_{relation}"
    sig_key = f"{state_key}_shape"

    use_auto = st.toggle("Auto-suggest scores (model)", value=True)

    shape_sig = (tuple(rows), tuple(cols))
    need_reset = (sig_key not in st.session_state) or (st.session_state[sig_key] != shape_sig)

    if need_reset:
        if use_auto:
            with st.spinner("Auto-scoring matrix (embeddings + optional NLI)…"):
                suggested, debug = auto_score_matrix(rows, cols)
            st.session_state[state_key] = suggested.copy()
            st.session_state[f"{state_key}_debug"] = debug
        else:
            st.session_state[state_key] = pd.DataFrame(0, index=rows, columns=cols)

        st.session_state[sig_key] = shape_sig

    if use_auto and f"{state_key}_debug" in st.session_state:
        with st.expander("Show auto-scoring debug (sim / NLI probs)", expanded=False):
            st.dataframe(st.session_state[f"{state_key}_debug"], use_container_width=True, height=260)

    current = st.session_state[state_key].copy()

    # ---- Human override editor with dropdown values -3..3 ----
    allowed_vals = [-3, -2, -1, 0, 1, 2, 3]

    # Use st.data_editor with selectbox columns
    col_cfg = {
        col: st.column_config.SelectboxColumn(
            col,
            options=allowed_vals,
            required=True
        )
        for col in current.columns
    }

    edited = st.data_editor(
        current,
        use_container_width=True,
        height=520,
        column_config=col_cfg
    )

    # Save back
    st.session_state[state_key] = edited.copy()

    # ---------------------------
    # 5) Export (your formatting)
    # ---------------------------
    st.markdown("## 5) Export to Excel")


    def export_matrix_to_excel_bytes(df_matrix, rotate_headers=55, bold_last_goal=True, sheet_name="Matrix"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        for r in dataframe_to_rows(df_matrix, index=False, header=True):
            ws.append(r)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Column widths
        ws.column_dimensions["A"].width = 55
        for col in range(2, df_matrix.shape[1] + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # Header row formatting
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        for col in range(2, df_matrix.shape[1] + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                textRotation=rotate_headers,
                horizontal="center",
                vertical="bottom",
                wrap_text=True
            )

        # Row heights
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 60 if row == 1 else 55

        # Borders + alignment
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.border = border
                if row > 1 and col == 1:
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                elif row > 1 and col > 1:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # Optional bold last column
        if bold_last_goal:
            last_col = ws.max_column
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=last_col).font = Font(bold=True)

        ws.freeze_panes = "B2"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    # Build export df: first column = row label, rest = matrix values
    df_export = edited.reset_index()
    df_export.rename(columns={df_export.columns[0]: row_label}, inplace=True)

    excel_bytes = export_matrix_to_excel_bytes(
        df_export,
        rotate_headers=55,
        bold_last_goal=True,
        sheet_name="Matrix"
    )

    st.download_button(
        "⬇️ Download coherence Excel (formatted matrix)",
        data=excel_bytes,
        file_name="gea_coherence_matrix.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================================
# ROUTER (uses your sidebar variable `page`)
# ============================================================
if page == "Home":
    render_home()
elif page == "Webscrapper":
    render_webscrapper()
elif page == "PDF Extractor":
    render_pdf_extractor()
elif page == "Coherence":
    render_coherence()


